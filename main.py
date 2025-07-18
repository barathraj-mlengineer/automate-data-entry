import streamlit as st
import pandas as pd
import time
import threading
import tempfile
import queue

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

from openpyxl import load_workbook

# Global flags
stop_flag = False
last_submitted_row = 0


def read_excel_preserving_format(uploaded_file):
    """Read Excel and return DataFrame with values displayed exactly as in Excel (e.g., 10% not 0.1)."""
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(uploaded_file.read())
        tmp_path = tmp.name

    wb = load_workbook(tmp_path, data_only=False)
    ws = wb.active

    data = []
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    for row in ws.iter_rows(min_row=2):
        values = []
        for cell in row:
            if cell.number_format in ['0%', '0.00%', '0.000%']:
                try:
                    raw_value = cell.value
                    if raw_value is not None:
                        formatted = f"{raw_value * 100:.0f}%"
                        values.append(formatted)
                    else:
                        values.append("")
                except:
                    values.append(str(cell.value))
            else:
                values.append(str(cell.value) if cell.value is not None else "")
        data.append(values)

    df = pd.DataFrame(data, columns=headers)
    return df.fillna("")


def submit_google_form(form_url, data_row, delay):
    try:
        options = Options()
        options.binary_location = "/usr/bin/google-chrome"
        options.add_argument("--headless")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--window-size=1920x1080")




        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
        driver.get(form_url)
        time.sleep(2)

        if "form" not in driver.page_source.lower():
            raise Exception("Google Form did not load properly.")

        input_elements = driver.find_elements(By.CSS_SELECTOR, "input[type='text']")
        if not input_elements:
            raise Exception("No input fields found in the form.")

        for idx, input_elem in enumerate(input_elements):
            if idx < len(data_row):
                value = str(data_row[idx])
                input_elem.send_keys(value)
                time.sleep(0.3)

        submit_button = driver.find_element(By.XPATH, '//span[contains(text(),"Submit") or contains(text(),"submit")]')
        submit_button.click()
        time.sleep(1.5)

        driver.quit()
        time.sleep(delay)
        return True

    except Exception as e:
        return f"❌ Error submitting form: {e}"


def automation_process(df, form_url, delay, start_row, end_row, status_queue):
    global stop_flag, last_submitted_row

    for index in range(start_row - 1, end_row):
        if stop_flag:
            status_queue.put(("warning", "🛑 Automation stopped by user."))
            break

        status_queue.put(("info", f"⏳ Submitting Row {index + 1} ..."))
        row = df.iloc[index]

        result = submit_google_form(form_url, row.tolist(), delay)
        if result is True:
            last_submitted_row = index + 1
            status_queue.put(("success", f"✅ Row {index + 1} submitted successfully."))
        else:
            status_queue.put(("error", f"{result}"))
            break

    if not stop_flag and last_submitted_row > 0:
        status_queue.put(("balloons", None))
        status_queue.put(("success", f"🎉 Finished. Last submitted row: {last_submitted_row}"))
    elif last_submitted_row == 0:
        status_queue.put(("warning", "⚠️ No rows were submitted."))


def main():
    global stop_flag, last_submitted_row
    st.set_page_config(page_title="📤 Excel to Google Form Bot", layout="centered")
    st.title("📤 Excel to Google Form Automation")
    st.markdown("Submit Excel rows to Google Form exactly as written, no formatting changes.")

    form_url = st.text_input("🔗 Enter Google Form URL")
    uploaded_file = st.file_uploader("📁 Upload Excel File (.xlsx only)", type=["xlsx"])
    delay = st.number_input("⏳ Delay between rows (in seconds)", min_value=1, value=30)

    if uploaded_file:
        try:
            df = read_excel_preserving_format(uploaded_file)
            total_rows = len(df)
            st.success(f"✅ Loaded {total_rows} rows")

            start_row = st.number_input("🔢 Start Row", min_value=1, max_value=total_rows, value=1)
            end_row = st.number_input("🔚 End Row", min_value=start_row, max_value=total_rows, value=total_rows)

            if "status_queue" not in st.session_state:
                st.session_state.status_queue = queue.Queue()

            start_button = st.button("🚀 Start Automation")
            stop_button = st.button("🛑 Stop Automation")

            if start_button:
                stop_flag = False
                last_submitted_row = 0
                thread = threading.Thread(target=automation_process,
                                          args=(df, form_url, delay, start_row, end_row, st.session_state.status_queue))
                thread.start()

            if stop_button:
                stop_flag = True
                st.warning("🛑 Stop signal sent.")
                st.info(f"⏹ Last submitted row: {last_submitted_row}")

            # Poll queue for messages
            while not st.session_state.status_queue.empty():
                level, message = st.session_state.status_queue.get()
                if level == "success":
                    st.success(message)
                elif level == "info":
                    st.info(message)
                elif level == "warning":
                    st.warning(message)
                elif level == "error":
                    st.error(message)
                elif level == "balloons":
                    st.balloons()

        except Exception as e:
            st.error(f"❌ Failed to read Excel file: {e}")
    else:
        st.info("📌 Upload a valid Excel file to begin.")


if __name__ == "__main__":
    main()
